#!/usr/bin/env python3
"""Build the master model spreadsheet from registry/models.yaml.

    python scripts/build_model_spreadsheet.py

Emits, all from the single sources of truth (registry/models.yaml for the
model-level view, registry/parameters.yaml for the parameter-level one):

    docs/MODEL_SPREADSHEET.xlsx   eight sheets (see below)
    docs/MODEL_SPREADSHEET.csv    the Models sheet, flat, for anyone
    docs/MODEL_PARAMETERS.csv     the Parameters sheet, flat

Sheets
    Models            one row per model, everything we know about it
    Coverage          12 diseases x 3 questions, so gaps read at a glance
    Key features      which predictor actually drives each model, measured
    Parameters        one row per PREDICTOR: coefficient, the transform applied
                      before the coefficient touches it, the effect size the
                      paper reports, and the measured swing beside it
    Axis differences  per disease, what changes between prediction, response
                      and prognosis, computed, not asserted
    Gaps              open cells and the candidates found for them
    Summary           the counts, with their denominators explained
    Glossary          what every column means

The predecessor of this script hard-coded its rows, which is how it came to
claim BCRAT was unverified months after it had been verified. Nothing here is
typed by hand.
"""

from __future__ import annotations

import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

import yaml

ROOT = Path(__file__).resolve().parents[1]
REGISTRY = ROOT / "registry" / "models.yaml"
PARAMETERS = ROOT / "registry" / "parameters.yaml"
DOCS = ROOT / "docs"
sys.path.insert(0, str(ROOT / "src"))
# Explicit rather than relying on Python putting the script's own directory on
# sys.path: that only happens when this is run AS a script, and the renderer
# test imports some of these modules instead.
sys.path.insert(0, str(Path(__file__).resolve().parent))

from cancerverse_baseline.reporting import (  # noqa: E402
    AXES,
    AXIS_LABEL,
    DISEASE_LABEL,
    SPREADSHEET_EXTRAS,
    build_rows,
)
from cancerverse_baseline.reporting import (
    COLUMNS as CORE_COLUMNS,
)
from reproducible_office import BUILD_EPOCH, make_reproducible  # noqa: E402

#: The shared table, then the spreadsheet-only extras.
COLUMNS = CORE_COLUMNS + SPREADSHEET_EXTRAS

# Candidates found 2026-08-06 on riskcalc.org. See docs/CANDIDATE_MODELS.md.
RISKCALC = "https://github.com/ClevelandClinicQHS/riskcalc-website/tree/main/"
CANDIDATES = {
    ("cvd", "response"): (
        None,
        "Derivable as a risk difference on PREVENT/SCORE2 x a trial relative risk "
        "reduction; no single published equation. A modelling decision, not a search.",
    ),
    ("ovarian", "response"): (
        None,
        "KELIM is a nonlinear population-PK model, not closed-form.",
    ),
}


# --------------------------------------------------------------------- data,
def load() -> list[dict]:
    return yaml.safe_load(REGISTRY.read_text())["models"]


def _flat(x) -> str:
    return " ".join(str(x or "").split())


def feature_table() -> tuple[dict, list[dict]]:
    """Run the sensitivity sweep. Returns (top predictor per model, all rows)."""
    try:
        import feature_importance as fi
    except Exception as exc:  # pragma: no cover
        print(f"  (feature sweep unavailable: {exc})")
        return {}, []

    top, rows = {}, []
    for model_id in list(fi.SPEC) + list(fi.CATEGORICAL_NOTE):
        try:
            res = fi.sweep(model_id)
        except Exception as exc:
            print(f"  (sweep failed for {model_id}: {exc})")
            continue
        if res is None:
            feat, why = fi.CATEGORICAL_NOTE.get(model_id, ("—", ""))
            top[model_id] = feat
            rows.append(
                {
                    "model": model_id,
                    "feature": feat,
                    "swing": None,
                    "share": None,
                    "unit": "",
                    "note": why,
                }
            )
            continue
        unit = "percentage points of risk" if fi.SPEC[model_id][3] == 100 else "points"
        if res["features"]:
            top[model_id] = res["features"][0]["feature"]
        for r in res["features"]:
            rows.append(
                {
                    "model": model_id,
                    "feature": r["feature"],
                    "swing": round(r["swing"], 3),
                    "share": r["share"],
                    "unit": unit,
                    "note": "",
                }
            )
    return top, rows


def parameter_rows(models: list[dict], feats: list[dict]) -> list[dict]:
    """One row per PARAMETER per model, from registry/parameters.yaml.

    Joins in the measured swing from the sensitivity sweep where the sweep
    covers that parameter, so the paper's claim about a predictor and our
    measurement of it sit on the same line and can disagree visibly. The join
    is fuzzy on purpose: the sweep labels features for humans ("age at
    menarche") while the registry keys them by our keyword argument
    ("age_menarche"), and forcing them to agree would mean renaming one to
    suit the other.
    """
    if not PARAMETERS.exists():  # pragma: no cover
        print("  (registry/parameters.yaml missing. Parameters sheet skipped)")
        return []
    data = yaml.safe_load(PARAMETERS.read_text())["parameters"]
    title_of = {m["id"]: _flat(m.get("title")) for m in models}
    disease_of = {
        m["id"]: DISEASE_LABEL.get(m.get("disease"), m.get("disease")) for m in models
    }

    swing: dict[tuple[str, str], dict] = {}
    for f in feats:
        swing[(f["model"], _norm(f["feature"]))] = f

    out = []
    for model_id, block in data.items():
        params = block.get("params", [])
        assigned = _assign_swings(swing, model_id, params)
        for i, p in enumerate(params):
            hit = assigned.get(i)
            out.append(
                {
                    "Model": title_of.get(model_id, model_id),
                    "Model id": model_id,
                    "Disease": disease_of.get(model_id, ""),
                    "Question": AXIS_LABEL.get(block.get("axis"), block.get("axis")),
                    "Parameter": p.get("name"),
                    "As the paper prints it": p.get("label"),
                    "Rank": p.get("importance_rank"),
                    "Coefficient": p.get("coefficient"),
                    "Points": _flat(p.get("points")) or "",
                    "Units": p.get("units") or "",
                    "Transform / preprocessing": _flat(p.get("transform")),
                    "Kind of transform": p.get("transform_kind") or "",
                    "Reference level": p.get("reference_level") or "",
                    "Effect size the paper reports": _flat(p.get("paper_effect")),
                    "Change in c-statistic (paper)": _flat(p.get("paper_delta_c")),
                    "Measured swing": (
                        round(hit["swing"], 3)
                        if hit and hit.get("swing") is not None
                        else None
                    ),
                    "Share of model swing": hit["share"] if hit else None,
                    "What will otherwise be misread": _flat(p.get("note")),
                }
            )
    return out


#: Tokens that carry no identifying information when matching a parameter name
#: to a sweep label. Units are here because the registry keys parameters by our
#: keyword argument, which encodes the unit (`bilirubin_umol_l`), while the
#: sweep labels them for a human ("bilirubin").
_NOISE = {
    "the",
    "a",
    "of",
    "per",
    "at",
    "in",
    "score",
    "level",
    "years",
    "year",
    "count",
    "status",
    "sex",
    "abnormal",
    "positive",
    "current",
    "total",
    "number",
    "no",
    "and",
    "class",
    "grade",
    "use",
    "history",
    # units
    "ml",
    "mg",
    "dl",
    "g",
    "l",
    "kg",
    "cm",
    "mm",
    "umol",
    "pmol",
    "mmol",
    "u",
    "ng",
    "bpm",
    "day",
    "lb",
    "hours",
    "decile",
    "yr",
}


def _tokens(s: str) -> set[str]:
    s = str(s or "").lower()
    for ch in "_-/()<>=%.,+":
        s = s.replace(ch, " ")
    return {w for w in s.split() if w and w not in _NOISE}


def _norm(s: str) -> str:
    """Loosest sensible key for matching a sweep label to a parameter name."""
    return " ".join(sorted(_tokens(s)))


#: (model id, parameter name) -> the sweep's label for it. Only for the few
#: pairs token matching cannot resolve, either because the two names share no
#: word ("glucose rise" vs `glucose_one_year_before_mg_dl`) or because several
#: parameters are equally good matches and the tie is correctly refused.
#: Everything not listed here matches on its own or is legitimately blank.
_SWEEP_ALIAS = {
    ("pbcg_extended", "famhist_1"): "first-degree family history",
    ("prevent", "sbp"): "systolic BP",
    ("prevent", "total_chol_mg_dl"): "total cholesterol",
    ("abc_method", "pepsinogen_i_ii_ratio"): "pepsinogen atrophy",
    ("msk_ovarian", "grade"): "grade 3",
    ("endpac", "glucose_one_year_before_mg_dl"): "glucose rise",
}


def _assign_swings(swing: dict, model_id: str, params: list[dict]) -> dict[int, dict]:
    """Match this model's parameters to swept features, ONE TO ONE.

    Returns {index into params: swing row}. A swept feature is claimed by at
    most one parameter and a parameter claims at most one feature.

    The one-to-one constraint is the whole point. A plain best-overlap match
    let `prior_psa` inherit PSA's swing, `famhist_2` inherit first-degree
    family history's, and `(age squared)` inherit age's, three numbers that
    would have been wrong in a way no reader could catch, because a measured
    swing looks like a measurement whatever produced it. Exact matches are
    assigned first, then the strongest remaining overlaps; an unresolved tie
    is left BLANK, which this workbook already defines as "not established".
    """
    features = {feat: row for (mid, feat), row in swing.items() if mid == model_id}
    if not features:
        return {}

    out: dict[int, dict] = {}
    taken: set[str] = set()

    # pass 0, the explicit alias table wins over anything inferred
    for i, p in enumerate(params):
        alias = _SWEEP_ALIAS.get((model_id, p.get("name")))
        if alias and _norm(alias) in features:
            out[i], _ = features[_norm(alias)], taken.add(_norm(alias))

    # pass 1, exact normalised equality, on either the keyword or the label
    for i, p in enumerate(params):
        if i in out:
            continue
        for candidate in (p.get("name"), p.get("label")):
            key = _norm(candidate)
            if key in features and key not in taken:
                out[i], _ = features[key], taken.add(key)
                break

    # pass 2, best token overlap among what is left
    scored = []
    for i, p in enumerate(params):
        if i in out:
            continue
        ptok = _tokens(p.get("name")) | _tokens(p.get("label"))
        for feat in features:
            n = len(ptok & set(feat.split()))
            if n:
                scored.append((n, i, feat))

    # A feature whose best claim is a tie is BURNED, not handed down to a
    # weaker claimant. Without this, pbcg_extended's "first-degree family
    # history", tied between famhist_1 and famhist_bca at three shared
    # tokens, fell through to famhist_2 at two, and second-degree family
    # history silently displayed first-degree's measured swing.
    for feat in features:
        claims = [n for n, _, f in scored if f == feat]
        if claims and claims.count(max(claims)) > 1:
            taken.add(feat)

    for n, i, feat in sorted(scored, key=lambda t: -t[0]):
        if i in out or feat in taken:
            continue
        if n < max(m for m, _, f in scored if f == feat):
            continue  # not this feature's best claimant
        out[i], _ = features[feat], taken.add(feat)
    return out


#: Read on a disease where all three axes are filled and the parameters still
#: barely overlap. Keyed by disease id; only where there is something to say
#: that the computed columns cannot say for themselves.
AXIS_READING = {
    "liver": "The one disease where a single pair of labs runs through all "
    "three questions, and it is the same two numbers each time. aMAP "
    "(will they get HCC) embeds ALBI whole as one 0.48-weighted term; "
    "ALBI (how is the liver) IS those two labs; HAP (will TACE help) "
    "dichotomises them at 36 g/L and 17 umol/L. Same inputs, three "
    "different transforms of them: continuous inside a nested model, "
    "continuous, then thresholded.",
    "prostate": "Age and PSA appear on all three axes and mean something "
    "different each time. In detection PSA is log2 and continuous. "
    "In CAPRA it is banded into five point levels and everything "
    "above 30 ng/mL is one bucket. In the dutasteride response "
    "model it is splined, and the answer is a DIFFERENCE between "
    "two arms rather than a level.",
    "lung": "Detection and response/prognosis share nothing at all. "
    "PLCOm2012 asks who will develop lung cancer and is built from "
    "smoking history and demography; LIPI asks how someone with it "
    "will do, from two blood tests taken on the day. No cigarette "
    "count appears in LIPI and no neutrophil count in PLCOm2012.",
    "cvd": "Three axes, no parameter common to all three, because the "
    "response model does not take patient variables at all. It takes a "
    "baseline risk (from PREVENT or SCORE2) and an LDL reduction, so "
    "the patient enters only through the detection model's output.",
    "breast": "Three axes, nothing shared, and the reason is that the "
    "questions use different eras of information. BCRAT asks about "
    "a woman with no diagnosis (menarche, births, biopsies); PREDICT "
    "asks about a tumour (size, nodes, grade, ER, HER2); the "
    "response arm asks about treatments. Nothing that predicts "
    "getting breast cancer helps predict surviving it.",
    "colorectal": "Age is the only shared parameter, and even it is not really "
    "shared: in CRC-PRO it is the dominant predictor (+0.163 of "
    "c-statistic) and splined at 47/60/72; in MSK rectal it is "
    "absent from the RFS model entirely and splined at "
    "36/52/64/78.7 for OS only.",
}


def axis_rows(models: list[dict], audit: dict) -> list[dict]:
    """One row per disease: what actually changes across the three questions.

    Computed from the audit rather than asserted, because the interesting
    finding is a negative one and a hand-written table would be tempted to
    round it off: of the five diseases with all three axes filled, only two
    share ANY parameter across all three.
    """
    by_id = {m["id"]: m for m in models}
    cell: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for model_id, block in audit.items():
        disease = by_id[model_id]["disease"]
        cell[disease][block["axis"]].append((model_id, block))

    out = []
    for disease in sorted(cell, key=lambda d: DISEASE_LABEL.get(d, d)):
        axes = cell[disease]
        sets = {}
        for axis, entries in axes.items():
            names: set[str] = set()
            for _, block in entries:
                names |= {
                    p["name"]
                    for p in block["params"]
                    if not str(p["name"]).strip().startswith("(")
                }
            sets[axis] = names
        shared = sorted(set.intersection(*sets.values())) if len(sets) > 1 else []

        # `axes=axes` binds this iteration's value explicitly. The closure is
        # only ever called inside the same iteration, so late binding never
        # actually bit, but that made it correct by accident rather than by
        # construction, and storing `tops` for later would have broken it.
        def tops(axis, axes=axes):
            return ", ".join(
                f"{b['top_parameter']} ({mid})" for mid, b in axes.get(axis, [])
            )

        out.append(
            {
                "Disease": DISEASE_LABEL.get(disease, disease),
                "Questions covered": len(axes),
                "Models": sum(len(v) for v in axes.values()),
                "Prediction, parameter that carries it": tops("detection"),
                "Response, parameter that carries it": tops("response"),
                "Prognosis, parameter that carries it": tops("prognosis"),
                "Parameters shared by EVERY covered question": (
                    ", ".join(shared)
                    if shared
                    else ("nothing" if len(sets) > 1 else "—")
                ),
                "What changes across the questions": AXIS_READING.get(disease, ""),
            }
        )
    return out


def summary(models: list[dict], rows: list[dict]) -> list[tuple[str, object]]:
    impl = [m for m in models if m.get("status") == "implemented"]
    checked = [m for m in impl if m.get("parity_status") in {"checked", "matched"}]
    src = Counter(m.get("open_source") for m in impl)
    fams = Counter(r["Architecture"] for r in rows)
    filled = {(m["disease"], m["axis"]) for m in impl}

    # Derive the reachable denominator from the registry rather than hardcoding
    # it. It has already been wrong once (24 when it should have been 26).
    from cancerverse_baseline.registry.load import progress_report

    rep = progress_report()
    reachable = rep["n_cells_reachable"]
    unreachable = rep["n_cells_unreachable"]

    L: list[tuple[str, object]] = [
        ("COVERAGE", ""),
        (
            "Diseases with at least one model",
            f"{len({m['disease'] for m in impl})} / 12",
        ),
        ("Disease x question cells filled", f"{len(filled)} / 36 nominal"),
        ("...of the cells we believe are reachable", f"{len(filled)} / {reachable}"),
        ("Models implemented", len(impl)),
        ("", ""),
        ("VERIFICATION", ""),
        ("Verified against an independent source", f"{len(checked)} / {len(impl)}"),
        ("Verification rate", f"{len(checked) / len(impl):.0%}"),
        ("Every check re-runnable offline", "yes, fixtures are committed"),
        ("", ""),
        ("HOW HARD THEY WERE TO GET", ""),
        ("Had a runnable reference implementation", src["available"]),
        ("Had only a hosted web calculator", src["web_only"]),
        ("Had nothing but the paper", src["none"]),
        ("", ""),
        ("ARCHITECTURES", f"{len(fams)} families"),
    ]
    for fam, n in fams.most_common():
        L.append((f"    {fam}", n))
    L += [
        ("", ""),
        ("A NOTE ON THE DENOMINATOR", ""),
        (
            "Why not 36?",
            f"For {unreachable} of the 36 cells we have not found a published equation. "
            "That is a statement about our search, not about the literature: it was "
            "12 until 2026-08-06, when a scan of one website's published source turned "
            "up equations for four cells the earlier survey had written off. "
            "Correcting it moved progress DOWN, because the survey had been wrong. "
            "Each open cell records its own blocker in registry/models.yaml; see "
            "the Gaps sheet.",
        ),
    ]
    return L


# ------------------------------------------------------------------ writers,
def write_csv(rows: list[dict]) -> Path:
    p = DOCS / "MODEL_SPREADSHEET.csv"
    with p.open("w", newline="", encoding="utf-8-sig") as fh:
        # build_rows carries private "_"-prefixed keys for the renderers.
        w = csv.DictWriter(fh, fieldnames=COLUMNS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    return p


def write_parameters_csv(params: list[dict]) -> Path:
    """The Parameters sheet, flat, for anyone without Excel."""
    p = DOCS / "MODEL_PARAMETERS.csv"
    if not params:  # pragma: no cover
        return p
    with p.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=list(params[0]))
        w.writeheader()
        w.writerows(params)
    return p


def write_xlsx(models, rows, feats, top, params, axes_rows) -> Path:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    INK, ACCENT = "16211F", "0F6E5C"
    HEAD = Font(bold=True, color="FFFFFF", size=10.5)
    FILL = PatternFill("solid", fgColor=INK)
    SUB = PatternFill("solid", fgColor="E4F0EC")
    OKF = PatternFill("solid", fgColor="E3F2E1")
    NOF = PatternFill("solid", fgColor="FDECEA")
    GAPF = PatternFill("solid", fgColor="FDF3E0")
    THIN = Side(style="thin", color="D0D7DE")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(vertical="top", wrap_text=True)
    TOPL = Alignment(vertical="top")

    def header(ws, labels):
        ws.append(labels)
        for c in ws[1]:
            c.font, c.fill, c.alignment, c.border = HEAD, FILL, WRAP, BORDER
        ws.row_dimensions[1].height = 34

    def widths(ws, spec, default=18):
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = spec.get(
                ws.cell(1, i).value, default
            )

    wb = Workbook()

    # ---- Models -----------------------------------------------------------
    ws = wb.active
    ws.title = "Models"
    header(ws, COLUMNS)
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
    # the formula is a displayed equation; monospace it so its alignment holds
    fcol = COLUMNS.index("Core formula") + 1
    for row in ws.iter_rows(min_row=2, min_col=fcol, max_col=fcol):
        row[0].font = Font(name="Menlo", size=9)
    vcol = COLUMNS.index("Verified?") + 1
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment, c.border = WRAP, BORDER
        cell = row[vcol - 1]
        cell.fill = OKF if cell.value == "Yes" else NOF
        cell.font = Font(bold=True, color=ACCENT if cell.value == "Yes" else "9A3412")
    widths(
        ws,
        {
            "Disease": 15,
            "Question": 12,
            "Model": 40,
            "Architecture": 20,
            "Architecture detail": 60,
            "Core formula": 54,
            "Parameters": 11,
            "Verified?": 10,
            "How it was verified": 80,
            "Re-run the check": 62,
            "Reference data captured by": 40,
            "Fixture": 38,
            "Where the equation came from": 66,
            "Citation": 44,
            "Reference implementation available?": 28,
            "Top predictor": 24,
            "Scope / caveats": 50,
            "Our code": 44,
        },
    )
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions

    # ---- Coverage ---------------------------------------------------------
    ws2 = wb.create_sheet("Coverage")
    header(ws2, ["Disease", *AXIS_LABEL.values()])
    by = defaultdict(lambda: defaultdict(list))
    for r in rows:
        by[r["Disease"]][r["Question"]].append(r)
    for disease in sorted(DISEASE_LABEL.values()):
        line = [disease]
        for axis in AXIS_LABEL.values():
            entries = by.get(disease, {}).get(axis, [])
            if not entries:
                line.append("— no model")
            else:
                line.append(
                    "\n".join(
                        f"{e['Model']}\n[{'verified' if e['Verified?'] == 'Yes' else 'NOT verified'}]"
                        for e in entries
                    )
                )
        ws2.append(line)
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.alignment, c.border = WRAP, BORDER
            if isinstance(c.value, str) and c.value.startswith("—"):
                c.fill = GAPF
        row[0].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 18
    for col in "BCD":
        ws2.column_dimensions[col].width = 46
    for i in range(2, ws2.max_row + 1):
        ws2.row_dimensions[i].height = 62

    # ---- Key features -----------------------------------------------------
    ws3 = wb.create_sheet("Key features")
    header(
        ws3,
        [
            "Model",
            "Disease",
            "Predictor",
            "Swing across its clinical range",
            "Unit",
            "Share of this model's total swing",
            "Note",
        ],
    )
    dis_of = {r["Model"]: r["Disease"] for r in rows}
    id_to_title = {m["id"]: _flat(m.get("title")) for m in models}
    for f in feats:
        title = id_to_title.get(f["model"], f["model"])
        ws3.append(
            [
                title,
                dis_of.get(title, ""),
                f["feature"],
                f["swing"],
                f["unit"],
                f["share"],
                f["note"],
            ]
        )
    for row in ws3.iter_rows(min_row=2):
        for c in row:
            c.alignment, c.border = TOPL, BORDER
        row[3].number_format = "0.00"
        row[5].number_format = "0%"
        row[2].alignment = WRAP
        row[6].alignment = WRAP
    if ws3.max_row > 1:
        ws3.conditional_formatting.add(
            f"F2:F{ws3.max_row}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=1,
                color=ACCENT,
                showValue=True,
            ),
        )
    widths(
        ws3,
        {
            "Model": 40,
            "Disease": 15,
            "Predictor": 32,
            "Swing across its clinical range": 16,
            "Unit": 24,
            "Share of this model's total swing": 18,
            "Note": 70,
        },
    )
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions

    # ---- Parameters -------------------------------------------------------
    # One row per predictor per model: what it is, what the coefficient is,
    # what happens to the value before the coefficient touches it, and what
    # the paper said about how much it mattered.
    if params:
        ws7 = wb.create_sheet("Parameters")
        pcols = list(params[0])
        header(ws7, pcols)
        for r in params:
            ws7.append([r[c] for c in pcols])
        idx = {c: i for i, c in enumerate(pcols)}
        for row in ws7.iter_rows(min_row=2):
            for c in row:
                c.alignment, c.border = TOPL, BORDER
            for col in (
                "Transform / preprocessing",
                "Reference level",
                "Effect size the paper reports",
                "Change in c-statistic (paper)",
                "What will otherwise be misread",
                "Points",
            ):
                row[idx[col]].alignment = WRAP
            row[idx["Coefficient"]].number_format = "0.000000"
            row[idx["Measured swing"]].number_format = "0.00"
            row[idx["Share of model swing"]].number_format = "0%"
            # rank 1 is the parameter that carries the model, make it findable
            if row[idx["Rank"]].value == 1:
                for c in row:
                    c.fill = SUB
            # a note is a warning, not decoration
            if row[idx["What will otherwise be misread"]].value:
                row[idx["What will otherwise be misread"]].fill = GAPF
        widths(
            ws7,
            {
                "Model": 38,
                "Model id": 16,
                "Disease": 15,
                "Question": 11,
                "Parameter": 26,
                "As the paper prints it": 34,
                "Rank": 6,
                "Coefficient": 14,
                "Points": 30,
                "Units": 18,
                "Transform / preprocessing": 58,
                "Kind of transform": 22,
                "Reference level": 24,
                "Effect size the paper reports": 44,
                "Change in c-statistic (paper)": 26,
                "Measured swing": 12,
                "Share of model swing": 12,
                "What will otherwise be misread": 88,
            },
        )
        ws7.freeze_panes = "E2"
        ws7.auto_filter.ref = ws7.dimensions

    # ---- Axis differences -------------------------------------------------
    # "Is there any difference between prediction, response and prognosis?"
    # answered per disease, from the audit rather than from memory.
    if axes_rows:
        ws8 = wb.create_sheet("Axis differences")
        acols = list(axes_rows[0])
        header(ws8, acols)
        aidx = {c: i for i, c in enumerate(acols)}
        for r in axes_rows:
            ws8.append([r[c] for c in acols])
        for row in ws8.iter_rows(min_row=2):
            for c in row:
                c.alignment, c.border = WRAP, BORDER
            row[aidx["Questions covered"]].alignment = TOPL
            row[aidx["Models"]].alignment = TOPL
            shared = row[aidx["Parameters shared by EVERY covered question"]]
            if shared.value == "nothing":
                shared.fill = NOF
            elif shared.value not in ("—", ""):
                shared.fill = OKF
        widths(
            ws8,
            {
                "Disease": 17,
                "Questions covered": 9,
                "Models": 8,
                "Prediction, parameter that carries it": 32,
                "Response, parameter that carries it": 32,
                "Prognosis, parameter that carries it": 32,
                "Parameters shared by EVERY covered question": 26,
                "What changes across the questions": 96,
            },
        )
        ws8.freeze_panes = "B2"

    # ---- Gaps -------------------------------------------------------------
    ws4 = wb.create_sheet("Gaps")
    header(ws4, ["Disease", "Question", "Status", "Candidate found", "Notes", "Source"])
    impl_cells = {
        (m["disease"], m["axis"]) for m in models if m.get("status") == "implemented"
    }
    tier = {}
    for m in models:
        if m.get("status") in {"gap", "catalog"}:
            tier.setdefault((m.get("disease"), m.get("axis")), m.get("repro_tier"))
    for did in sorted(DISEASE_LABEL, key=lambda k: DISEASE_LABEL[k]):
        for axis in AXES:
            if (did, axis) in impl_cells:
                continue
            cand, note = CANDIDATES.get((did, axis), (None, ""))
            # No cell resolves to tier D any more (see model_table.GAP_CAPTION),
            # but keep the branch: if one ever does, it must say so loudly rather
            # than silently reading as an ordinary open cell.
            unreachable = tier.get((did, axis)) == "D"
            status = (
                "REGRESSION: marked never-published"
                if unreachable
                else "open, not implemented"
            )
            if cand:
                status += " · CANDIDATE FOUND"
            ws4.append(
                [
                    DISEASE_LABEL[did],
                    AXIS_LABEL[axis],
                    status,
                    cand or "",
                    note,
                    RISKCALC + cand.split(" ")[0] if cand else "",
                ]
            )
    for row in ws4.iter_rows(min_row=2):
        for c in row:
            c.alignment, c.border = WRAP, BORDER
        if "CANDIDATE" in str(row[2].value):
            for c in row:
                c.fill = SUB
    widths(
        ws4,
        {
            "Disease": 15,
            "Question": 12,
            "Status": 34,
            "Candidate found": 46,
            "Notes": 62,
            "Source": 60,
        },
    )
    ws4.freeze_panes = "A2"

    # ---- Summary ----------------------------------------------------------
    ws5 = wb.create_sheet("Summary")
    header(ws5, ["Metric", "Value"])
    for k, v in summary(models, rows):
        ws5.append([k, v])
        if k and not k.startswith("    ") and v == "":
            for c in ws5[ws5.max_row]:
                c.font, c.fill = Font(bold=True, color=ACCENT), SUB
    for row in ws5.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    ws5.column_dimensions["A"].width = 46
    ws5.column_dimensions["B"].width = 62

    # ---- Glossary ---------------------------------------------------------
    ws6 = wb.create_sheet("Glossary")
    header(ws6, ["Column", "What it means"])
    GLOSSARY = [
        ("Disease", "The disease the model is about, not the organ."),
        (
            "Question",
            "Prediction (will they get it?), Response (will treatment "
            "help?), Prognosis (what is the outlook?). These are the "
            "three questions every disease is asked.",
        ),
        (
            "Architecture",
            "What class of model it is. All of these are "
            "fixed-coefficient statistical models, a handful of "
            "numbers estimated once and printed in a paper. None "
            "is a neural network.",
        ),
        (
            "Core formula",
            "The equation, as it would be set in the paper. "
            "Caveats live in Scope / caveats, not here.",
        ),
        (
            "Year",
            "Publication year of the model, cross-checked against its own citation.",
        ),
        (
            "Discrimination (AUC / C-index)",
            "How well the model separates who has the outcome from who does not. "
            "1.0 is perfect, 0.5 is a coin flip; most clinical models sit at "
            "0.65-0.85. Blank where we have not read the number from the paper: "
            "it enters this table only from a source actually read, so a blank "
            "means we have not read it, not that the model was never evaluated.",
        ),
        (
            "Developed on",
            "The cohort the model was fitted on. Useful mainly "
            "for judging whether it transports to your patients.",
        ),
        (
            "Public repository",
            "A public repo containing the model's CODE, "
            "where one exists. A hosted calculator is not a "
            "repository and is not counted here.",
        ),
        (
            "Source",
            "The publication. Every model has one, and it is the thing "
            "to check us against.",
        ),
        (
            "Verified?",
            "Whether our output was compared against something we "
            "did not write. Not 'the tests pass', our own tests "
            "cannot tell us we read a paper correctly.",
        ),
        (
            "How it was verified",
            "Which of six routes was used. Full account "
            "with exact numbers in docs/VERIFICATION.md.",
        ),
        (
            "Re-run the check",
            "A command you can paste. The comparison runs "
            "offline against a committed fixture.",
        ),
        (
            "Top predictor",
            "Which input moves the output most, measured by "
            "sweeping each input across its clinical range. Not "
            "the largest coefficient, age often has a small "
            "coefficient and still dominates, because it ranges "
            "over fifty years.",
        ),
        (
            "How hard it was to get",
            "Whether the model arrived as runnable "
            "code, as a hosted calculator we had to "
            "reverse, or as a paper we had to "
            "transcribe.",
        ),
        (
            "Parameters",
            "Roughly how many fitted numbers the model carries, for "
            "contrast with modern architectures. A transformer "
            "carries 10^9 to 10^12.",
        ),
        (
            "Scope / caveats",
            "The population the model was built for. Applying "
            "one outside its scope produces a number that "
            "looks valid and is not.",
        ),
        ("", ""),
        ("Parameters sheet", ""),
        (
            "Rank",
            "1 is the parameter that carries the model. How that was "
            "decided differs by model and is recorded per model in "
            "registry/parameters.yaml under `importance_basis`, the "
            "paper's own ranking where it published one, the measured "
            "sweep otherwise. Rank-1 rows are shaded.",
        ),
        (
            "Coefficient",
            "The fitted number AS PUBLISHED. Blank where the model "
            "is a point score (see Points) or where the term is a "
            "spline/lookup rather than one slope. Comparing "
            "coefficients ACROSS models is meaningless; comparing "
            "them within one is only meaningful once Units and "
            "Transform are read.",
        ),
        (
            "Transform / preprocessing",
            "What happens to the value before the coefficient touches it. This is "
            "the column most likely to be skipped and most likely to cause a "
            "wrong number: PLCOm2012's smoking intensity enters as a RECIPROCAL, "
            "ALBI's bilirubin as log10 of a micromolar value, CRC-PRO's weight in "
            "POUNDS because the spline knots assume the app's own conversion.",
        ),
        (
            "Reference level",
            "For categorical predictors, the level that "
            "carries no term. Several models pick one a reader "
            "would not guess. MSK gastric's Lauren reference "
            "is diffuse (the worst type), MSK ovarian's "
            "residual-disease reference is the middle "
            "category, CRC-PRO's ethnicity reference is Black.",
        ),
        (
            "Effect size the paper reports",
            "The odds/hazard ratio the paper prints, verbatim where possible. "
            "Blank means the paper printed none, not that the effect is unknown.",
        ),
        (
            "Change in c-statistic (paper)",
            "Only CRC-PRO publishes this: how much each variable added to "
            "discrimination as it entered a forward-stepwise fit. It is the most "
            "direct importance evidence in the whole repository.",
        ),
        (
            "Measured swing",
            "How far the model's output moves when this input "
            "is swept across its clinical range, everything "
            "else held at a reference patient. Blank where the "
            "sweep does not cover that parameter, a blank is "
            "not a zero.",
        ),
        (
            "What will otherwise be misread",
            "Sign flips, non-monotone effects, unit traps, dead branches, "
            "reference levels that invert a reading. Shaded because each one is a "
            "way to get a wrong answer that still looks right.",
        ),
        ("", ""),
        (
            "A note on blanks",
            "Blank cells in this workbook mean we have not established the value, "
            "not that no value exists. That distinction matters here: four cells "
            "recorded as having no published equation turned out to have one.",
        ),
    ]
    for k, v in GLOSSARY:
        ws6.append([k, v])
    for row in ws6.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
        row[0].font = Font(bold=True)
    ws6.column_dimensions["A"].width = 32
    ws6.column_dimensions["B"].width = 96

    p = DOCS / "MODEL_SPREADSHEET.xlsx"
    # openpyxl stamps docProps/core.xml with the wall-clock time of the save,
    # to the second. Pinning it before saving is half of what makes this file a
    # function of the registry rather than of when it was built.
    wb.properties.created = wb.properties.modified = BUILD_EPOCH
    wb.save(p)
    make_reproducible(p)
    return p




def main() -> int:
    models = load()
    print("running the feature sweep...")
    top, feats = feature_table()
    rows = build_rows(models, top)
    params = parameter_rows(models, feats)
    audit = (
        yaml.safe_load(PARAMETERS.read_text())["parameters"]
        if PARAMETERS.exists()
        else {}
    )
    axes_rows = axis_rows(models, audit) if audit else []
    for p in (
        write_csv(rows),
        write_parameters_csv(params),
        write_xlsx(models, rows, feats, top, params, axes_rows),
    ):
        print(f"wrote {p.relative_to(ROOT)}")
    print(
        f"\n{len(rows)} implemented models, {len(feats)} feature rows, "
        f"{len(params)} parameter rows"
    )
    for k, v in summary(models, rows):
        if v != "" and not k.startswith("    ") and k.isupper() is False:
            print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
